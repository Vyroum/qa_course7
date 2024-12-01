from openpyxl import load_workbook

workbook = load_workbook("tmp/sample-1-sheet.xlsx")
sheet = workbook.active
print(sheet.cell(row=2, column=2).value)
print(sheet.max_column)
print(sheet.max_row)